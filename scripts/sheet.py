"""Writing a spreadsheet a person can actually open.

The content fields run to paragraphs and carry their own line breaks. A CSV
encodes that with quoting, and a reader that ignores the quoting shears every
following column onto the wrong row — which is what Google Sheets did. XLSX
carries the cell boundaries in the format itself, so there is nothing to
mis-detect.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

ROW_HEIGHT = 15
MAX_WIDTH = 55


def write_sheet(path: Path, frame: pd.DataFrame, sheet_name: str,
                lead: list[str] | None = None) -> None:
    """One sheet, frozen header, one line per row.

    `lead` names the columns to move to the front, so whatever the reader needs
    to decide is visible without scrolling.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    front = [c for c in (lead or []) if c in frame.columns]
    ordered = frame[front + [c for c in frame.columns if c not in front]]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        ordered.to_excel(writer, sheet_name=sheet_name, index=False)
        sheet = writer.sheets[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Never wrap. Wrapping makes every row as tall as its longest paragraph,
        # and one ingredients list turns a single product into half a screen.
        # Clipped text is still whole in the cell and shows in the formula bar.
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical="center")

        # Explicit, so a reader that thinks the content is too long does not
        # fall back to auto-fit.
        for n in range(2, sheet.max_row + 1):
            sheet.row_dimensions[n].height = ROW_HEIGHT

        for i, column in enumerate(ordered.columns, 1):
            longest = ordered[column].astype(str).str.len().max() if len(ordered) else 0
            width = min(max(12, int(longest or 0) + 2), MAX_WIDTH)
            sheet.column_dimensions[get_column_letter(i)].width = width
